"""
create_templates.py
-------------------
Generates Excel import templates for SiteLink with:
  - Drop-down lists for all categorical/lookup columns
  - Data validation for numeric fields (lat, long, azimuth, etc.)
  - Province/ward/RNC/antenna data fetched from the PostgreSQL database
  - Formatted headers matching the import parser column names

Usage:
    python create_templates.py
    # → writes template_site.xlsx, template_cell_3g.xlsx,
    #           template_cell_4g.xlsx, template_cell_5g.xlsx,
    #           template_antenna.xlsx  into ./templates/

Requirements:
    pip install openpyxl psycopg2-binary python-dotenv
"""

from __future__ import annotations

import os
import sys
from typing import Any, Dict, List, Optional, Sequence, Tuple

# ── optional .env loading ─────────────────────────────────────────────────────
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

# ══════════════════════════════════════════════════════════════════════════════
# 1.  DATABASE HELPERS
# ══════════════════════════════════════════════════════════════════════════════

DB_PARAMS = {
    "host":     "localhost",
    "port":     "5432",
    "dbname":   "sitelink_db",
    "user":     "sitelink",
    "password": "sitelink_pass",
}

_DB_AVAILABLE = False
_db_conn      = None


def _get_conn():
    global _db_conn, _DB_AVAILABLE
    if _db_conn is not None:
        return _db_conn
    try:
        import psycopg2
        _db_conn     = psycopg2.connect(**DB_PARAMS)
        _DB_AVAILABLE = True
        print("[DB] Connected to PostgreSQL successfully.")
        return _db_conn
    except Exception as exc:
        print(f"[DB] Cannot connect ({exc}). Templates will use static fallback values.")
        _DB_AVAILABLE = False
        return None


def _query(sql: str, params=None) -> List[tuple]:
    conn = _get_conn()
    if conn is None:
        return []
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params or ())
            return cur.fetchall()
    except Exception as exc:
        print(f"[DB] Query error: {exc}")
        try:
            conn.rollback()
        except Exception:
            pass
        return []


# ── Lookup loaders ────────────────────────────────────────────────────────────

def load_tinh_list() -> List[str]:
    rows = _query(
        "SELECT DISTINCT ten_tinh FROM dropdown_tinh_xa_phuong "
        "WHERE ten_tinh IS NOT NULL ORDER BY ten_tinh"
    )
    result = [r[0] for r in rows if r[0]]
    if not result:
        # static fallback – a representative subset
        result = [
            "An Giang", "Bà Rịa - Vũng Tàu", "Bắc Giang", "Bắc Kạn",
            "Bạc Liêu", "Bắc Ninh", "Bến Tre", "Bình Định", "Bình Dương",
            "Bình Phước", "Bình Thuận", "Cà Mau", "Cần Thơ", "Cao Bằng",
            "Đà Nẵng", "Đắk Lắk", "Đắk Nông", "Điện Biên", "Đồng Nai",
            "Đồng Tháp", "Gia Lai", "Hà Giang", "Hà Nam", "Hà Nội",
            "Hà Tĩnh", "Hải Dương", "Hải Phòng", "Hậu Giang", "Hòa Bình",
            "Hưng Yên", "Khánh Hòa", "Kiên Giang", "Kon Tum", "Lai Châu",
            "Lâm Đồng", "Lạng Sơn", "Lào Cai", "Long An", "Nam Định",
            "Nghệ An", "Ninh Bình", "Ninh Thuận", "Phú Thọ", "Phú Yên",
            "Quảng Bình", "Quảng Nam", "Quảng Ngãi", "Quảng Ninh",
            "Quảng Trị", "Sóc Trăng", "Sơn La", "Tây Ninh", "Thái Bình",
            "Thái Nguyên", "Thanh Hóa", "Thừa Thiên Huế", "Tiền Giang",
            "TP. Hồ Chí Minh", "Trà Vinh", "Tuyên Quang", "Vĩnh Long",
            "Vĩnh Phúc", "Yên Bái",
        ]
    return result


def load_phuong_xa_for_tinh(tinh: str) -> List[str]:
    rows = _query(
        "SELECT DISTINCT ten_phuong_xa FROM dropdown_tinh_xa_phuong "
        "WHERE ten_tinh = %s AND ten_phuong_xa IS NOT NULL "
        "ORDER BY ten_phuong_xa",
        (tinh,)
    )
    return [r[0] for r in rows if r[0]]


def load_all_phuong_xa() -> List[str]:
    """Flat list of all unique wards (for the hidden lookup sheet)."""
    rows = _query(
        "SELECT DISTINCT ten_phuong_xa FROM dropdown_tinh_xa_phuong "
        "WHERE ten_phuong_xa IS NOT NULL ORDER BY ten_phuong_xa LIMIT 5000"
    )
    result = [r[0] for r in rows if r[0]]
    if not result:
        result = ["Phường 1", "Phường 2", "Xã An Lạc", "Xã Bình Hưng"]
    return result


def load_rnc_names() -> Dict[str, List[str]]:
    """Returns {vendor: [rnc_name, ...]} dict."""
    rows = _query(
        "SELECT vendor, name FROM rnc_names ORDER BY vendor, name"
    )
    result: Dict[str, List[str]] = {}
    for vendor, name in rows:
        result.setdefault(vendor, []).append(name)

    if not result:
        # static fallback matching main.py _RNC_DATA
        result = {
            "Ericsson": [
                "RHNCG1E","RHNCG2E","RHNCG3E","RHNCG4E","RHNHM1E","RHNHM2E",
                "RHNHM3E","RQNHL1E","RQNHL2E","RSG103E","RSG011E","RSG072E",
                "RSG091E","RSG092E","RSG093E","RSG094E","RSG095E","RSG097E",
                "RSG104E","RSG105E","RSGBC2E","RSGBI2E","RSGBT2E","RSGHM1E",
                "RSGTB2E",
            ],
            "Huawei": [
                "iHNCG1H","iHNCG3H","iHNHM1H","iHNHM2H","iHNHM3H","iHNHM5H",
                "iHNHM6H","iHNHM7H","iHNHM8H","RHNCG1H","RHNCG3H","RHNCG4H",
                "RHNHM3H","RHNHM4H","RHNHM5H","RHNHM6H","RHNHM7H","RHNHM8H",
                "RDNG01H","RDNG02H","RQBDH1H","RCTCR11H","RCTCR12H",
            ],
            "Nokia": [
                "RDNCL3N","RDNCL4N","RDNCL5N","RDNCL7N","RDNCL8N","RDNCL9N",
                "RDNST10N","RDNST12N","RDNST13N","RDNST14N","RDNST15N",
                "RDNST7N","RCTCR1N","RCTCR2N","RCTCR8N","RDNBH5N","RSG091N",
                "RSG092N","RSG093N","RSG094N","RSG095N","RSG097N","RSG098N",
                "RSG099N","RSG105N","RTGMT3N",
            ],
            "ZTE": ["RCTCR3Z","RCTCR4Z","RCTCR5Z","RCTCR6Z"],
        }
    return result


def load_antenna_names() -> List[str]:
    rows = _query("SELECT name FROM antennas ORDER BY name LIMIT 2000")
    result = [r[0] for r in rows if r[0]]
    if not result:
        result = ["CHƯA XÁC ĐỊNH", "Antenna_A", "Antenna_B"]
    return result


def load_phan_loai_tram() -> List[str]:
    rows = _query(
        "SELECT value FROM dropdown_general WHERE category = 'phan_loai_tram' ORDER BY value"
    )
    result = [r[0] for r in rows if r[0]]
    if not result:
        result = ["IBC", "Macro outdoor", "IBC + Outdoor", "Smallcell", "miniDAS"]
    return result


# ══════════════════════════════════════════════════════════════════════════════
# 2.  STYLING CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════

HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
HDR_FONT  = Font(color="FFFFFF", bold=True, size=10)
REQ_FILL  = PatternFill("solid", fgColor="FFF2CC")   # yellow – required
OPT_FILL  = PatternFill("solid", fgColor="DDEEFF")   # light blue – optional
LOCK_FILL = PatternFill("solid", fgColor="F0F0F0")   # grey – read-only hint
NOTE_FILL = PatternFill("solid", fgColor="E2EFDA")   # green – notes row

THIN  = Side(style="thin",   color="B0B0B0")
THICK = Side(style="medium", color="1F4E79")
BORDER_CELL  = Border(left=THIN,  right=THIN,  top=THIN,  bottom=THIN)
BORDER_HDR   = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)
CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

LOOKUP_SHEET = "_Lookups"   # hidden sheet that holds long list values


# ══════════════════════════════════════════════════════════════════════════════
# 3.  LOW-LEVEL HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _style_header_row(ws: Worksheet, n_cols: int, row: int = 1) -> None:
    ws.row_dimensions[row].height = 32
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = CENTER
        cell.border    = BORDER_HDR


def _style_note_row(ws: Worksheet, notes: List[str], row: int = 2) -> None:
    ws.row_dimensions[row].height = 45
    for col, note in enumerate(notes, start=1):
        cell = ws.cell(row=row, column=col, value=note)
        cell.fill      = NOTE_FILL
        cell.font      = Font(italic=True, size=8, color="555555")
        cell.alignment = LEFT
        cell.border    = BORDER_CELL


def _style_data_rows(ws: Worksheet, n_cols: int,
                     start_row: int = 3, end_row: int = 1002,
                     required_cols: Optional[set] = None) -> None:
    required_cols = required_cols or set()
    for row in range(start_row, end_row + 1):
        alt = (row % 2 == 0)
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            if col in required_cols:
                cell.fill = REQ_FILL
            elif alt:
                cell.fill = PatternFill("solid", fgColor="F7FBFF")
            cell.alignment = LEFT
            cell.border    = BORDER_CELL


def _set_col_width(ws: Worksheet, col: int, width: float) -> None:
    ws.column_dimensions[get_column_letter(col)].width = width


def _freeze(ws: Worksheet, cell: str = "A3") -> None:
    ws.freeze_panes = cell


def _add_autofilter(ws: Worksheet, n_cols: int) -> None:
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"


# ── Lookup sheet helpers ──────────────────────────────────────────────────────

def _ensure_lookup_sheet(wb: Workbook) -> Worksheet:
    if LOOKUP_SHEET in wb.sheetnames:
        return wb[LOOKUP_SHEET]
    ws = wb.create_sheet(LOOKUP_SHEET)
    ws.sheet_state = "hidden"
    return ws


def _write_lookup_col(wb: Workbook, col_idx: int,
                      values: List[str], header: str) -> str:
    """
    Write *values* into column *col_idx* of the hidden lookup sheet.
    Returns an Excel formula string like  _Lookups!$B$2:$B$64
    """
    ws  = _ensure_lookup_sheet(wb)
    col = get_column_letter(col_idx)
    ws.cell(row=1, column=col_idx, value=header).font = Font(bold=True)
    for i, v in enumerate(values, start=2):
        ws.cell(row=i, column=col_idx, value=v)
    end_row = len(values) + 1
    return f"{quote_sheetname(LOOKUP_SHEET)}!${col}$2:${col}${end_row}"


def _dv_list_formula(formula: str,
                     first_data_row: int = 3,
                     last_data_row:  int = 1002) -> DataValidation:
    """DataValidation using a named range / lookup-sheet reference."""
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,   # False = show the arrow
        showErrorMessage=True,
        errorTitle="Giá trị không hợp lệ",
        error="Vui lòng chọn từ danh sách",
    )
    return dv


def _dv_list_inline(values: Sequence[str],
                    first_data_row: int = 3,
                    last_data_row:  int = 1002) -> DataValidation:
    """DataValidation with inline comma-separated values (max ~255 chars)."""
    formula = '"' + ",".join(values) + '"'
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Giá trị không hợp lệ",
        error="Vui lòng chọn từ danh sách",
    )
    return dv


def _dv_decimal(min_val: float, max_val: float,
                title: str = "Giá trị không hợp lệ",
                error: str = "") -> DataValidation:
    dv = DataValidation(
        type="decimal",
        operator="between",
        formula1=str(min_val),
        formula2=str(max_val),
        allow_blank=True,
        showErrorMessage=True,
        errorTitle=title,
        error=error or f"Phải trong khoảng {min_val} – {max_val}",
    )
    return dv


def _dv_whole(min_val: int, max_val: int,
              title: str = "Giá trị không hợp lệ",
              error: str = "") -> DataValidation:
    dv = DataValidation(
        type="whole",
        operator="between",
        formula1=str(min_val),
        formula2=str(max_val),
        allow_blank=True,
        showErrorMessage=True,
        errorTitle=title,
        error=error or f"Phải là số nguyên trong khoảng {min_val} – {max_val}",
    )
    return dv


def _apply_dv(ws: Worksheet, dv: DataValidation,
              col: int, first_row: int = 3, last_row: int = 1002) -> None:
    col_letter = get_column_letter(col)
    dv.sqref   = f"{col_letter}{first_row}:{col_letter}{last_row}"
    ws.add_data_validation(dv)


def _col_map(columns: List[Tuple]) -> Dict[str, int]:
    """Build {header_name: col_index_1based} from columns list."""
    return {col[0]: idx + 1 for idx, col in enumerate(columns)}


# ══════════════════════════════════════════════════════════════════════════════
# 4.  LOOKUP DATA  (loaded once, shared across templates)
# ══════════════════════════════════════════════════════════════════════════════

print("Loading lookup data from database …")
TINH_LIST      = load_tinh_list()
ALL_PHUONG_XA  = load_all_phuong_xa()
RNC_GROUPED    = load_rnc_names()
ALL_RNC        = sorted({n for names in RNC_GROUPED.values() for n in names})
ANTENNA_NAMES  = load_antenna_names()
PHAN_LOAI      = load_phan_loai_tram()

# Static constant lists (same as web form)
MIEN_LIST      = ["MB", "MT", "MN"]
VENDOR_LIST    = ["Ericsson", "Nokia", "Huawei", "ZTE", "Samsung"]
MORAN_LIST     = ["VNPT HOST", "MBF HOST"]
MIMO_LIST      = ["2x2", "4x4", "8x8"]
VUNG_LIST      = ["Indoor", "Outdoor"]
CELL_VIP_LIST  = ["VIP", "VVIP"]
SITE_VIP_LIST  = ["VIP", "VVIP"]
BOOL_LIST      = ["x", ""]          # "x" = True, blank = False
CHUNG_3G       = ["3G", "3G/4G", "2G/3G/4G", "3G/4G/5G", "3G/5G"]
CHUNG_4G       = ["4G", "2G/4G", "3G/4G", "2G/3G/4G", "4G/5G"]
MU_MIMO_LIST   = ["Yes", "No"]

# Validation bounds (matching validators.ts and import_excel.py)
VN_LAT_MIN, VN_LAT_MAX  = 8.33,   23.39
VN_LON_MIN, VN_LON_MAX  = 102.14, 109.47
AZI_MIN,    AZI_MAX      = 0,      359

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "templates")
os.makedirs(OUTPUT_DIR, exist_ok=True)


# ══════════════════════════════════════════════════════════════════════════════
# 5.  TEMPLATE: SITE
# ══════════════════════════════════════════════════════════════════════════════

def create_site_template() -> None:
    """
    Columns match parse_site_excel() and SiteFormPage.tsx / SiteBase schema.
    Column order follows the export_sites() function for familiarity.
    """

    # (header, note, width)
    columns: List[Tuple[str, str, float]] = [
        ("Mien",             "MB / MT / MN – bắt buộc",                         8),
        ("Tinh",             "Tỉnh / Thành phố – bắt buộc, chọn từ danh sách",  28),
        ("Phuong xa",        "Phường / Xã – chọn từ danh sách",                  28),
        ("Site name (cu)",   "Tên site cũ (nếu đổi tên)",                        24),
        ("Site name",        "Tên site hiện tại – bắt buộc, duy nhất",           28),
        ("Site VIP",         "VIP / VVIP – để trống nếu không phải",             12),
        ("Ma PTM",           "Mã PTM",                                            16),
        ("Lat",              "Latitude 8.33 – 23.39",                            14),
        ("Long",             "Longitude 102.14 – 109.47",                        14),
        ("Tram 2G",          "x = có, để trống = không",                         10),
        ("Tram 3G",          "x = có, để trống = không",                         10),
        ("Tram 4G",          "x = có, để trống = không",                         10),
        ("Tram 5G",          "x = có, để trống = không",                         10),
        ("Repeater",         "x = có, để trống = không",                         10),
        ("Booster",          "x = có, để trống = không",                         10),
        ("Node truyen dan only", "x = có, để trống = không",                     20),
        ("Tram phu song TSCA",   "x = có, để trống = không",                     18),
        ("Phan loai tram",   "Chọn từ danh sách",                                22),
        ("MORAN 3G",         "VNPT HOST / MBF HOST",                             18),
        ("MORAN 4G",         "VNPT HOST / MBF HOST",                             18),
        ("MORAN 5G",         "VNPT HOST / MBF HOST",                             18),
        ("Do cao dinh cot anten", "Độ cao đỉnh cột anten (m)",                   22),
        ("Do cao cot anten", "Độ cao cột anten mặt đất (m)",                     20),
        ("Dia chi",          "Địa chỉ chi tiết",                                 30),
        ("Ghi chu",          "Ghi chú",                                          30),
    ]

    wb  = Workbook()
    ws  = wb.active
    ws.title = "Sites"

    n_cols       = len(columns)
    FIRST_DATA   = 3
    LAST_DATA    = 1002
    required_set = {1, 2, 5}   # Mien(1), Tinh(2), Site name(5) – 1-based

    # ── headers & notes ───────────────────────────────────────────────────────
    for idx, (hdr, note, width) in enumerate(columns, start=1):
        ws.cell(row=1, column=idx, value=hdr)
        ws.cell(row=2, column=idx, value=note)
        _set_col_width(ws, idx, width)

    _style_header_row(ws, n_cols, row=1)
    _style_note_row(ws, [c[1] for c in columns], row=2)
    _style_data_rows(ws, n_cols, FIRST_DATA, LAST_DATA, required_set)
    _freeze(ws, "A3")
    _add_autofilter(ws, n_cols)

    cm = _col_map(columns)

    # ── Lookup sheet columns ──────────────────────────────────────────────────
    lk_tinh_ref   = _write_lookup_col(wb, 1, TINH_LIST,     "Tinh")
    lk_xa_ref     = _write_lookup_col(wb, 2, ALL_PHUONG_XA, "PhuongXa")
    lk_phanloai   = _write_lookup_col(wb, 3, PHAN_LOAI,     "PhanLoai")

    # ── Data validations ──────────────────────────────────────────────────────

    # Mien
    dv_mien = _dv_list_inline(MIEN_LIST)
    _apply_dv(ws, dv_mien, cm["Mien"], FIRST_DATA, LAST_DATA)

    # Tinh – lookup sheet
    dv_tinh = _dv_list_formula(lk_tinh_ref)
    _apply_dv(ws, dv_tinh, cm["Tinh"], FIRST_DATA, LAST_DATA)

    # Phuong xa – lookup sheet (all wards; per-province cascade not possible
    # in plain xlsx without VBA, so we allow the full list)
    dv_xa = _dv_list_formula(lk_xa_ref)
    _apply_dv(ws, dv_xa, cm["Phuong xa"], FIRST_DATA, LAST_DATA)

    # Site VIP
    dv_svip = _dv_list_inline(SITE_VIP_LIST)
    _apply_dv(ws, dv_svip, cm["Site VIP"], FIRST_DATA, LAST_DATA)

    # Lat / Long
    dv_lat = _dv_decimal(VN_LAT_MIN, VN_LAT_MAX,
                         "Latitude không hợp lệ",
                         f"Latitude phải trong khoảng {VN_LAT_MIN} – {VN_LAT_MAX}")
    _apply_dv(ws, dv_lat, cm["Lat"], FIRST_DATA, LAST_DATA)

    dv_lon = _dv_decimal(VN_LON_MIN, VN_LON_MAX,
                         "Longitude không hợp lệ",
                         f"Longitude phải trong khoảng {VN_LON_MIN} – {VN_LON_MAX}")
    _apply_dv(ws, dv_lon, cm["Long"], FIRST_DATA, LAST_DATA)

    # Boolean columns
    bool_cols = [
        "Tram 2G","Tram 3G","Tram 4G","Tram 5G",
        "Repeater","Booster","Node truyen dan only","Tram phu song TSCA",
    ]
    dv_bool = _dv_list_inline(BOOL_LIST)
    for col_name in bool_cols:
        dv = _dv_list_inline(BOOL_LIST)
        _apply_dv(ws, dv, cm[col_name], FIRST_DATA, LAST_DATA)

    # Phan loai tram
    dv_pl = _dv_list_formula(lk_phanloai)
    _apply_dv(ws, dv_pl, cm["Phan loai tram"], FIRST_DATA, LAST_DATA)

    # MORAN
    for col_name in ["MORAN 3G", "MORAN 4G", "MORAN 5G"]:
        dv = _dv_list_inline(MORAN_LIST)
        _apply_dv(ws, dv, cm[col_name], FIRST_DATA, LAST_DATA)

    # Height fields – positive decimal
    for col_name in ["Do cao dinh cot anten", "Do cao cot anten"]:
        dv = _dv_decimal(0, 999, "Độ cao không hợp lệ", "Phải là số dương (m)")
        _apply_dv(ws, dv, cm[col_name], FIRST_DATA, LAST_DATA)

    # ── Sample row ────────────────────────────────────────────────────────────
    sample = {
        "Mien": "MB", "Tinh": TINH_LIST[0] if TINH_LIST else "Hà Nội",
        "Site name": "HNI_XXXX_001", "Ma PTM": "PTM001",
        "Lat": 21.0285, "Long": 105.8542,
        "Tram 4G": "x", "Phan loai tram": PHAN_LOAI[0] if PHAN_LOAI else "Macro outdoor",
    }
    for col_name, val in sample.items():
        if col_name in cm:
            ws.cell(row=FIRST_DATA, column=cm[col_name], value=val)

    path = os.path.join(OUTPUT_DIR, "template_site.xlsx")
    wb.save(path)
    print(f"  ✓  {path}")


# ══════════════════════════════════════════════════════════════════════════════
# 6.  SHARED CELL COLUMN BUILDER
# ══════════════════════════════════════════════════════════════════════════════

_COMMON_CELL_COLS: List[Tuple[str, str, float]] = [
    ("Mien",            "MB / MT / MN",                            8),
    ("Tinh",            "Tỉnh / Thành phố – chọn từ danh sách",   28),
    ("Phuong xa",       "Phường / Xã – chọn từ danh sách",         28),
    ("Site Name",       "Tên site – bắt buộc",                     28),
    ("Site Name Old",   "Tên site cũ (nếu có)",                    24),
    ("Cell Name",       "Tên cell – bắt buộc",                     28),
    ("Cell Name Old",   "Tên cell cũ (nếu có)",                    24),
    ("Cell VIP",        "VIP / VVIP",                              10),
    ("MORAN",           "VNPT HOST / MBF HOST",                    18),
    ("Lat",             "Latitude 8.33 – 23.39",                   14),
    ("Long",            "Longitude 102.14 – 109.47",               14),
    ("Vung phu song",   "Indoor / Outdoor",                        14),
    ("Vendor",          "Chọn từ danh sách",                       14),
    ("Do cao anten",    "Độ cao anten (m), số dương",              16),
    ("Azimuth",         "Góc phương vị 0 – 359",                   12),
    ("M-tilt",          "Mechanical tilt",                         10),
    ("E-Tilt",          "Electrical tilt",                         10),
    ("Total Tilt",      "M-tilt + E-Tilt",                         12),
    ("Loai Anten",      "Chọn từ danh sách antenna",               35),
    ("Baseband",        "Tên thiết bị baseband",                   18),
    ("RF",              "Tên thiết bị RF",                         16),
    ("Cell ID",         "Cell ID (chuỗi / số)",                    14),
    ("MIMO",            "2x2 / 4x4 / 8x8",                        10),
    ("Cell max power (dBm)", "Công suất tối đa cell (dBm)",        20),
    ("BBUname",         "Tên BBU",                                 16),
    ("Cell status (at dump time)", "Trạng thái cell",              26),
]


def _apply_common_cell_validations(
    wb: Workbook,
    ws: Worksheet,
    cm: Dict[str, int],
    lookup_col_offset: int = 1,     # starting column in _Lookups for this template
) -> int:
    """
    Applies all common cell validations.
    Returns the next free lookup column index.
    """
    FIRST_DATA, LAST_DATA = 3, 1002
    lc = lookup_col_offset

    # Tinh
    lk_tinh = _write_lookup_col(wb, lc, TINH_LIST, "Tinh"); lc += 1
    dv = _dv_list_formula(lk_tinh)
    _apply_dv(ws, dv, cm["Tinh"], FIRST_DATA, LAST_DATA)

    # Phuong xa
    lk_xa = _write_lookup_col(wb, lc, ALL_PHUONG_XA, "PhuongXa"); lc += 1
    dv = _dv_list_formula(lk_xa)
    _apply_dv(ws, dv, cm["Phuong xa"], FIRST_DATA, LAST_DATA)

    # Mien
    dv = _dv_list_inline(MIEN_LIST)
    _apply_dv(ws, dv, cm["Mien"], FIRST_DATA, LAST_DATA)

    # Cell VIP
    dv = _dv_list_inline(CELL_VIP_LIST)
    _apply_dv(ws, dv, cm["Cell VIP"], FIRST_DATA, LAST_DATA)

    # MORAN
    dv = _dv_list_inline(MORAN_LIST)
    _apply_dv(ws, dv, cm["MORAN"], FIRST_DATA, LAST_DATA)

    # Lat / Long
    dv = _dv_decimal(VN_LAT_MIN, VN_LAT_MAX, "Latitude không hợp lệ",
                     f"Latitude phải trong khoảng {VN_LAT_MIN} – {VN_LAT_MAX}")
    _apply_dv(ws, dv, cm["Lat"], FIRST_DATA, LAST_DATA)

    dv = _dv_decimal(VN_LON_MIN, VN_LON_MAX, "Longitude không hợp lệ",
                     f"Longitude phải trong khoảng {VN_LON_MIN} – {VN_LON_MAX}")
    _apply_dv(ws, dv, cm["Long"], FIRST_DATA, LAST_DATA)

    # Vung phu song
    dv = _dv_list_inline(VUNG_LIST)
    _apply_dv(ws, dv, cm["Vung phu song"], FIRST_DATA, LAST_DATA)

    # Vendor
    dv = _dv_list_inline(VENDOR_LIST)
    _apply_dv(ws, dv, cm["Vendor"], FIRST_DATA, LAST_DATA)

    # Do cao anten
    dv = _dv_decimal(0, 999, "Độ cao không hợp lệ", "Phải là số dương (m)")
    _apply_dv(ws, dv, cm["Do cao anten"], FIRST_DATA, LAST_DATA)

    # Azimuth
    dv = _dv_whole(AZI_MIN, AZI_MAX, "Azimuth không hợp lệ",
                   f"Azimuth phải trong khoảng {AZI_MIN} – {AZI_MAX}")
    _apply_dv(ws, dv, cm["Azimuth"], FIRST_DATA, LAST_DATA)

    # Tilt fields – allow any decimal (no strict bounds in code, just numeric)
    for col_name in ["M-tilt", "E-Tilt", "Total Tilt"]:
        dv = _dv_decimal(-30, 30, f"{col_name} không hợp lệ",
                         f"{col_name} thường trong khoảng -30 đến 30")
        _apply_dv(ws, dv, cm[col_name], FIRST_DATA, LAST_DATA)

    # Loai Anten – lookup sheet
    lk_ant = _write_lookup_col(wb, lc, ANTENNA_NAMES, "LoaiAnten"); lc += 1
    dv = _dv_list_formula(lk_ant)
    _apply_dv(ws, dv, cm["Loai Anten"], FIRST_DATA, LAST_DATA)

    # MIMO
    dv = _dv_list_inline(MIMO_LIST)
    _apply_dv(ws, dv, cm["MIMO"], FIRST_DATA, LAST_DATA)

    return lc   # next free lookup col


def _build_cell_wb(
    sheet_title: str,
    extra_cols: List[Tuple[str, str, float]],
    required_col_names: Optional[set] = None,
) -> Tuple[Workbook, Worksheet, Dict[str, int]]:
    """Create a workbook with common + extra columns, styled."""
    columns  = _COMMON_CELL_COLS + extra_cols
    required = required_col_names or {"Site Name", "Cell Name"}

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    n_cols     = len(columns)
    FIRST_DATA = 3
    LAST_DATA  = 1002
    req_idx    = {idx+1 for idx, (h, _, _) in enumerate(columns) if h in required}

    for idx, (hdr, note, width) in enumerate(columns, start=1):
        ws.cell(row=1, column=idx, value=hdr)
        ws.cell(row=2, column=idx, value=note)
        _set_col_width(ws, idx, width)

    _style_header_row(ws, n_cols, row=1)
    _style_note_row(ws, [c[1] for c in columns], row=2)
    _style_data_rows(ws, n_cols, FIRST_DATA, LAST_DATA, req_idx)
    _freeze(ws, "A3")
    _add_autofilter(ws, n_cols)

    cm = _col_map(columns)
    return wb, ws, cm


# ══════════════════════════════════════════════════════════════════════════════
# 7.  TEMPLATE: CELL 3G
# ══════════════════════════════════════════════════════════════════════════════

def create_cell3g_template() -> None:
    """
    Extra 3G-specific columns after common block.
    Matching Cell3GBase schema and parse_cell3g_excel() extra_fields.
    """
    extra_cols: List[Tuple[str, str, float]] = [
        ("Chung anten",         "3G / 3G/4G / 2G/3G/4G / 3G/4G/5G / 3G/5G",     20),
        ("RNC Name",            "Tên RNC – chọn theo Vendor",                      18),
        ("ARFCN",               "ARFCN (chuỗi/số)",                               12),
        ("UARFCN",              "UARFCN (chuỗi/số)",                              12),
        ("LAC",                 "Location Area Code",                              12),
        ("RAC",                 "Routing Area Code",                               12),
        ("PSC",                 "Primary Scrambling Code",                         12),
        ("URAId",               "URA ID",                                          10),
        ("CPICH power (dBm)",   "CPICH power (dBm)",                              18),
    ]

    wb, ws, cm = _build_cell_wb("Cell_3G", extra_cols)
    FIRST_DATA, LAST_DATA = 3, 1002

    # Common validations (uses lookup cols 1, 2, 3)
    next_lc = _apply_common_cell_validations(wb, ws, cm, lookup_col_offset=1)

    # Chung anten 3G
    dv = _dv_list_inline(CHUNG_3G)
    _apply_dv(ws, dv, cm["Chung anten"], FIRST_DATA, LAST_DATA)

    # RNC Name – all vendors pooled into lookup sheet
    lk_rnc = _write_lookup_col(wb, next_lc, ALL_RNC, "RNCName"); next_lc += 1
    dv = _dv_list_formula(lk_rnc)
    _apply_dv(ws, dv, cm["RNC Name"], FIRST_DATA, LAST_DATA)

    # CPICH power – numeric range (-30 to 50 dBm typical)
    dv = _dv_decimal(-30, 50, "CPICH power không hợp lệ",
                     "CPICH power thường trong khoảng -30 đến 50 dBm")
    _apply_dv(ws, dv, cm["CPICH power (dBm)"], FIRST_DATA, LAST_DATA)

    # Cell max power
    dv = _dv_decimal(-30, 50, "Cell max power không hợp lệ",
                     "Cell max power thường trong khoảng -30 đến 50 dBm")
    _apply_dv(ws, dv, cm["Cell max power (dBm)"], FIRST_DATA, LAST_DATA)

    # Sample row
    sample = {
        "Mien": "MB",
        "Tinh": TINH_LIST[0] if TINH_LIST else "Hà Nội",
        "Site Name": "HNI_XXXX_001",
        "Cell Name": "HNI_XXXX_001_C1",
        "Vendor": "Huawei",
        "Azimuth": 120,
        "MIMO": "2x2",
        "Chung anten": "3G",
    }
    for k, v in sample.items():
        if k in cm:
            ws.cell(row=FIRST_DATA, column=cm[k], value=v)

    path = os.path.join(OUTPUT_DIR, "template_cell_3g.xlsx")
    wb.save(path)
    print(f"  ✓  {path}")


# ══════════════════════════════════════════════════════════════════════════════
# 8.  TEMPLATE: CELL 4G
# ══════════════════════════════════════════════════════════════════════════════

def create_cell4g_template() -> None:
    """
    Extra 4G-specific columns matching Cell4GBase schema
    and parse_cell4g_excel() extra_fields.
    """
    extra_cols: List[Tuple[str, str, float]] = [
        ("Chung anten",      "4G / 2G/4G / 3G/4G / 2G/3G/4G / 4G/5G",    20),
        ("EnodeB ID",        "eNodeB ID",                                   16),
        ("EARFCN",           "E-UTRA Absolute Radio Frequency Channel",     14),
        ("TAC",              "Tracking Area Code",                          12),
        ("PCI",              "Physical Cell Identity 0-503",                12),
        ("Root Sequence ID", "Root Sequence Index",                         18),
        ("Bandwitdh",        "Bandwidth (MHz) – vd: 5, 10, 15, 20",        16),
        ("ECI",              "E-UTRAN Cell Identifier",                     16),
    ]

    wb, ws, cm = _build_cell_wb("Cell_4G", extra_cols)
    FIRST_DATA, LAST_DATA = 3, 1002

    next_lc = _apply_common_cell_validations(wb, ws, cm, lookup_col_offset=1)

    # Chung anten 4G
    dv = _dv_list_inline(CHUNG_4G)
    _apply_dv(ws, dv, cm["Chung anten"], FIRST_DATA, LAST_DATA)

    # PCI 0-503
    dv = _dv_whole(0, 503, "PCI không hợp lệ", "PCI phải trong khoảng 0 – 503")
    _apply_dv(ws, dv, cm["PCI"], FIRST_DATA, LAST_DATA)

    # Cell max power
    dv = _dv_decimal(-30, 50, "Cell max power không hợp lệ",
                     "Cell max power thường trong khoảng -30 đến 50 dBm")
    _apply_dv(ws, dv, cm["Cell max power (dBm)"], FIRST_DATA, LAST_DATA)

    # Sample row
    sample = {
        "Mien": "MN",
        "Tinh": TINH_LIST[-1] if TINH_LIST else "TP. Hồ Chí Minh",
        "Site Name": "HCM_XXXX_001",
        "Cell Name": "HCM_XXXX_001_C1",
        "Vendor": "Ericsson",
        "Azimuth": 0,
        "MIMO": "4x4",
        "Chung anten": "4G",
        "PCI": 100,
    }
    for k, v in sample.items():
        if k in cm:
            ws.cell(row=FIRST_DATA, column=cm[k], value=v)

    path = os.path.join(OUTPUT_DIR, "template_cell_4g.xlsx")
    wb.save(path)
    print(f"  ✓  {path}")


# ══════════════════════════════════════════════════════════════════════════════
# 9.  TEMPLATE: CELL 5G
# ══════════════════════════════════════════════════════════════════════════════

def create_cell5g_template() -> None:
    """
    Extra 5G-specific columns matching Cell5GBase schema
    and parse_cell5g_excel() extra_fields.
    Note: 5G does NOT have chung_anten field.
    """
    extra_cols: List[Tuple[str, str, float]] = [
        ("gNodeB ID",        "gNodeB ID",                              16),
        ("TAC",              "Tracking Area Code",                     12),
        ("PCI",              "Physical Cell Identity 0-1007",          12),
        ("Root Sequence ID", "Root Sequence Index",                    18),
        ("SSB-ARFCN",        "SSB Absolute Radio Frequency Channel",   14),
        ("Center-ARFCN",     "Center Frequency ARFCN",                 16),
        ("GSCN",             "Global Synchronization Channel Number",  14),
        ("Bandwidth (MHz)",  "Bandwidth (MHz) – vd: 50, 100, 200",    16),
        ("NCI",              "NR Cell Identity",                       16),
        ("MU-MIMO",          "Yes / No",                               12),
    ]

    wb, ws, cm = _build_cell_wb("Cell_5G", extra_cols)
    FIRST_DATA, LAST_DATA = 3, 1002

    next_lc = _apply_common_cell_validations(wb, ws, cm, lookup_col_offset=1)

    # PCI 0-1007 for 5G (NR spec)
    dv = _dv_whole(0, 1007, "PCI không hợp lệ", "NR PCI phải trong khoảng 0 – 1007")
    _apply_dv(ws, dv, cm["PCI"], FIRST_DATA, LAST_DATA)

    # MU-MIMO
    dv = _dv_list_inline(MU_MIMO_LIST)
    _apply_dv(ws, dv, cm["MU-MIMO"], FIRST_DATA, LAST_DATA)

    # Cell max power
    dv = _dv_decimal(-30, 60, "Cell max power không hợp lệ",
                     "Cell max power thường trong khoảng -30 đến 60 dBm")
    _apply_dv(ws, dv, cm["Cell max power (dBm)"], FIRST_DATA, LAST_DATA)

    # Sample row
    sample = {
        "Mien": "MT",
        "Tinh": TINH_LIST[10] if len(TINH_LIST) > 10 else "Đà Nẵng",
        "Site Name": "DNG_XXXX_001",
        "Cell Name": "DNG_XXXX_001_C1_5G",
        "Vendor": "Nokia",
        "Azimuth": 240,
        "MIMO": "8x8",
        "MU-MIMO": "Yes",
        "PCI": 200,
    }
    for k, v in sample.items():
        if k in cm:
            ws.cell(row=FIRST_DATA, column=cm[k], value=v)

    path = os.path.join(OUTPUT_DIR, "template_cell_5g.xlsx")
    wb.save(path)
    print(f"  ✓  {path}")


# ══════════════════════════════════════════════════════════════════════════════
# 10. TEMPLATE: ANTENNA
# ══════════════════════════════════════════════════════════════════════════════

def create_antenna_template() -> None:
    """
    Matching AntennaBase schema and import_antenna_excel() parser.
    """
    columns: List[Tuple[str, str, float]] = [
        ("Name",            "Tên antenna – bắt buộc, duy nhất",               35),
        ("Band",            "Băng tần – vd: 900, 1800, 900-1800-2100",        22),
        ("5G_AAU",          "x = là 5G AAU, để trống = không",                10),
        ("No_of_ports",     "Số cổng (số nguyên dương)",                      14),
        ("No_of_beam",      "Số beam (số nguyên dương)",                      14),
        ("Horizontal BW",   "Horizontal beamwidth – vd: 65°",                 16),
        ("Vertical BW",     "Vertical beamwidth – vd: 7°",                    14),
        ("Gain",            "Gain (dBi) – vd: 17.5",                          12),
        ("Etilt",           "Electrical tilt range – vd: 0-10",               14),
        ("H",               "Height (mm)",                                     10),
        ("W",               "Width (mm)",                                      10),
        ("D",               "Depth (mm)",                                      10),
        ("Weight",          "Weight (kg)",                                     10),
        ("Connector type",  "Loại đầu nối – vd: 4.3-10, 7/16",               18),
        ("Ghi chú",         "Ghi chú",                                        30),
    ]

    wb  = Workbook()
    ws  = wb.active
    ws.title = "Antennas"

    n_cols     = len(columns)
    FIRST_DATA = 3
    LAST_DATA  = 1002
    required_set = {1}   # Name is required

    for idx, (hdr, note, width) in enumerate(columns, start=1):
        ws.cell(row=1, column=idx, value=hdr)
        ws.cell(row=2, column=idx, value=note)
        _set_col_width(ws, idx, width)

    _style_header_row(ws, n_cols, row=1)
    _style_note_row(ws, [c[1] for c in columns], row=2)
    _style_data_rows(ws, n_cols, FIRST_DATA, LAST_DATA, required_set)
    _freeze(ws, "A3")
    _add_autofilter(ws, n_cols)

    cm = _col_map(columns)

    # ── Data validations ──────────────────────────────────────────────────────

    # 5G_AAU boolean
    dv = _dv_list_inline(BOOL_LIST)
    _apply_dv(ws, dv, cm["5G_AAU"], FIRST_DATA, LAST_DATA)

    # No_of_ports – positive integer
    dv = _dv_whole(1, 32, "Số cổng không hợp lệ", "Phải là số nguyên 1 – 32")
    _apply_dv(ws, dv, cm["No_of_ports"], FIRST_DATA, LAST_DATA)

    # No_of_beam – positive integer
    dv = _dv_whole(1, 64, "Số beam không hợp lệ", "Phải là số nguyên 1 – 64")
    _apply_dv(ws, dv, cm["No_of_beam"], FIRST_DATA, LAST_DATA)

    # Band – common values as dropdown (also allows free text)
    common_bands = [
        "700", "850", "900", "1800", "2100", "2600", "3500",
        "900-1800", "900-2100", "1800-2100",
        "700-1800-2100", "900-1800-2100",
        "700-1800-2100-2600", "3500-26000",
    ]
    dv = _dv_list_inline(common_bands)
    _apply_dv(ws, dv, cm["Band"], FIRST_DATA, LAST_DATA)

    # Common connector types
    connector_types = ["4.3-10", "7/16 DIN", "N-Type", "SMA", "TNC", "EIA 7/8\""]
    dv = _dv_list_inline(connector_types)
    _apply_dv(ws, dv, cm["Connector type"], FIRST_DATA, LAST_DATA)

    # Sample row
    sample = {
        "Name":           "HUAWEI_AAU5613-1",
        "Band":           "2100",
        "5G_AAU":         "",
        "No_of_ports":    4,
        "No_of_beam":     1,
        "Horizontal BW":  "65°",
        "Vertical BW":    "7°",
        "Gain":           "17.5",
        "Etilt":          "0-10",
        "H":              "1340",
        "W":              "385",
        "D":              "177",
        "Weight":         "17",
        "Connector type": "4.3-10",
    }
    for k, v in sample.items():
        if k in cm:
            ws.cell(row=FIRST_DATA, column=cm[k], value=v)

    path = os.path.join(OUTPUT_DIR, "template_antenna.xlsx")
    wb.save(path)
    print(f"  ✓  {path}")


# ══════════════════════════════════════════════════════════════════════════════
# 11. LEGEND SHEET  (added to every workbook for user guidance)
# ══════════════════════════════════════════════════════════════════════════════

def _add_legend_sheet(wb: Workbook, tech: str = "") -> None:
    ws = wb.create_sheet("Hướng dẫn", 0)   # insert at beginning

    # Make it the first visible sheet
    rows = [
        ("HƯỚNG DẪN SỬ DỤNG FILE TEMPLATE SITELINK", None),
        (None, None),
        ("MÀU SẮC", "Ý NGHĨA"),
        ("Nền VÀNG",    "Cột bắt buộc – phải có dữ liệu"),
        ("Nền XANH",    "Cột tùy chọn – có thể để trống"),
        ("Nền XANH LÁ (dòng 2)", "Ghi chú / hướng dẫn từng cột"),
        (None, None),
        ("QUY TẮC NHẬP LIỆU", None),
        ("Cột có drop-down",   "Chỉ chọn từ danh sách – KHÔNG tự nhập tự do"),
        ("Cột boolean (x/blank)", "Nhập chữ x (thường) để bật, để trống để tắt"),
        ("Cột Lat/Long",    "Phải trong phạm vi lãnh thổ Việt Nam"),
        ("Azimuth",         "Phải trong khoảng 0 – 359"),
        ("Tilt",            "Thường trong khoảng -30 đến 30"),
        (None, None),
        ("QUY TẮC IMPORT", None),
        ("Cột CÓ trong file + ô TRỐNG",   "→ Xóa dữ liệu trường đó"),
        ("Cột KHÔNG CÓ trong file",        "→ Giữ nguyên dữ liệu hiện tại trong DB"),
        ("Dòng 1",  "Header – tên cột (KHÔNG sửa)"),
        ("Dòng 2",  "Ghi chú (KHÔNG sửa)"),
        ("Dòng 3+", "Dữ liệu – điền từ dòng 3"),
        (None, None),
        ("THÔNG TIN FILE", None),
        (f"Công nghệ", tech or "Site / Cell / Antenna"),
        ("Phiên bản", "SiteLink v1.2"),
    ]

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 48
    ws.row_dimensions[1].height     = 30

    title_font  = Font(bold=True, size=13, color="1F4E79")
    sect_font   = Font(bold=True, size=10, color="333333")
    hdr_fill2   = PatternFill("solid", fgColor="D9E1F2")

    for r_idx, (col_a, col_b) in enumerate(rows, start=1):
        ca = ws.cell(row=r_idx, column=1, value=col_a)
        cb = ws.cell(row=r_idx, column=2, value=col_b)
        ca.alignment = LEFT
        cb.alignment = LEFT

        if r_idx == 1:
            ca.font = title_font
        elif col_b is None and col_a and col_a.isupper():
            ca.font = sect_font
            ca.fill = hdr_fill2
            cb.fill = hdr_fill2
        elif col_a in ("MÀU SẮC", "QUY TẮC NHẬP LIỆU",
                       "QUY TẮC IMPORT", "THÔNG TIN FILE"):
            ca.font = sect_font

    # Color legend cells
    for row_label, fill_color in [
        ("Nền VÀNG",    "FFF2CC"),
        ("Nền XANH",    "DDEEFF"),
        ("Nền XANH LÁ (dòng 2)", "E2EFDA"),
    ]:
        for r_idx in range(1, len(rows) + 1):
            cell = ws.cell(row=r_idx, column=1)
            if cell.value == row_label:
                cell.fill = PatternFill("solid", fgColor=fill_color)
                break


# ══════════════════════════════════════════════════════════════════════════════
# 12. MAIN – generate all templates
# ══════════════════════════════════════════════════════════════════════════════

def _post_process(path: str, tech: str = "") -> None:
    """Open the saved workbook, add legend sheet, re-save."""
    wb = openpyxl.load_workbook(path)
    _add_legend_sheet(wb, tech)
    # Ensure _Lookups is last and hidden
    if LOOKUP_SHEET in wb.sheetnames:
        wb.move_sheet(LOOKUP_SHEET, offset=len(wb.sheetnames))
        wb[LOOKUP_SHEET].sheet_state = "hidden"
    wb.save(path)


if __name__ == "__main__":
    print(f"\n{'='*60}")
    print("  SiteLink – Excel Template Generator")
    print(f"{'='*60}\n")
    print(f"Output directory: {OUTPUT_DIR}\n")

    print("Generating template_site.xlsx …")
    create_site_template()
    _post_process(os.path.join(OUTPUT_DIR, "template_site.xlsx"), "Site")

    print("Generating template_cell_3g.xlsx …")
    create_cell3g_template()
    _post_process(os.path.join(OUTPUT_DIR, "template_cell_3g.xlsx"), "Cell 3G")

    print("Generating template_cell_4g.xlsx …")
    create_cell4g_template()
    _post_process(os.path.join(OUTPUT_DIR, "template_cell_4g.xlsx"), "Cell 4G")

    print("Generating template_cell_5g.xlsx …")
    create_cell5g_template()
    _post_process(os.path.join(OUTPUT_DIR, "template_cell_5g.xlsx"), "Cell 5G")

    print("Generating template_antenna.xlsx …")
    create_antenna_template()
    _post_process(os.path.join(OUTPUT_DIR, "template_antenna.xlsx"), "Antenna")

    print(f"\n{'='*60}")
    print("  All templates generated successfully!")
    print(f"{'='*60}\n")
    if not _DB_AVAILABLE:
        print("⚠️  WARNING: Database was not available.")
        print("   Province/ward/RNC/antenna lists used static fallback values.")
        print("   For full data, run with correct DB credentials set in .env\n")