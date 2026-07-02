"""
export.py
---------
Export endpoints for Sites, Cells 3G/4G/5G, and Antennas.
Returns Excel (.xlsx) files using openpyxl via a StreamingResponse.

All endpoints require authentication.
Token can be passed as:
  - Authorization: Bearer <token>  (normal axios call)
  - ?token=<token>                 (window.open download)
"""
from __future__ import annotations

import io
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.db.session import get_db
from app.models.site import Site
from app.models.cell_3g import Cell3G
from app.models.cell_4g import Cell4G
from app.models.cell_5g import Cell5G
from app.models.antenna import Antenna
from app.models.user import User
from app.utils.deps import get_current_user
from app.core.security import decode_access_token

router = APIRouter()

# ── Styling helpers ───────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left",   vertical="center")
THIN        = Side(style="thin", color="D0D0D0")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALT_FILL    = PatternFill("solid", fgColor="EBF3FB")


def _make_wb(headers: list[tuple[str, int]]) -> tuple[openpyxl.Workbook, object]:
    """Create a workbook with styled header row. Returns (wb, ws)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    return wb, ws


def _style_row(ws, row_idx: int, num_cols: int, alternate: bool):
    fill = ALT_FILL if alternate else None
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.alignment = LEFT
        cell.border    = BORDER
        if fill:
            cell.fill = fill


def _stream(wb: openpyxl.Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.read()]),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        },
    )


def _resolve_user(
    token: Optional[str],
    current_user: Optional[User],
    db: Session,
) -> User:
    """
    Support both:
      - Normal Bearer token (current_user already resolved by dependency)
      - ?token= query param (for window.open downloads)
    """
    if current_user:
        return current_user
    if token:
        payload = decode_access_token(token)
        if payload:
            user = db.query(User).filter(
                User.username == payload.get("sub")
            ).first()
            if user and user.is_active:
                return user
    raise HTTPException(status_code=401, detail="Not authenticated")


# ── optional auth dependency (allows token param fallback) ────────────────────
from fastapi.security import OAuth2PasswordBearer
from fastapi import Request

oauth2_optional = OAuth2PasswordBearer(
    tokenUrl="/api/v1/auth/login", auto_error=False
)


def get_optional_user(
    token_header: Optional[str] = Depends(oauth2_optional),
    token_param:  Optional[str] = Query(None, alias="token"),
    db: Session = Depends(get_db),
) -> User:
    # Try Bearer header first, then ?token= param
    raw = token_header or token_param
    if not raw:
        raise HTTPException(status_code=401, detail="Not authenticated")
    payload = decode_access_token(raw)
    if not payload:
        raise HTTPException(status_code=401, detail="Invalid token")
    user = db.query(User).filter(
        User.username == payload.get("sub")
    ).first()
    if not user or not user.is_active:
        raise HTTPException(status_code=401, detail="User inactive")
    return user


# ─────────────────────────────────────────────────────────────────────────────
# SITES export
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/sites")
def export_sites(
    search:  Optional[str]  = Query(None),
    mien:    Optional[str]  = Query(None),
    tinh:    Optional[str]  = Query(None),
    tram_3g: Optional[bool] = Query(None),
    tram_4g: Optional[bool] = Query(None),
    tram_5g: Optional[bool] = Query(None),
    db:      Session        = Depends(get_db),
    _:       User           = Depends(get_optional_user),
):
    q = db.query(Site)
    if search:  q = q.filter(Site.site_name.ilike(f"%{search}%"))
    if mien:    q = q.filter(Site.mien == mien)
    if tinh:    q = q.filter(Site.tinh == tinh)
    if tram_3g is not None: q = q.filter(Site.tram_3g == tram_3g)
    if tram_4g is not None: q = q.filter(Site.tram_4g == tram_4g)
    if tram_5g is not None: q = q.filter(Site.tram_5g == tram_5g)
    sites = q.order_by(Site.mien, Site.tinh, Site.site_name).all()

    headers = [
        ("STT",                        6),
        ("Mien",                       8),
        ("Tinh",                      22),
        ("Phuong xa",                 22),
        ("Site name (cu)",            22),
        ("Site name",                 25),
        ("Site VIP",                  10),
        ("Lat",                       14),
        ("Long",                      14),
        ("Tram 2G",                   10),
        ("Tram 3G",                   10),
        ("Tram 4G",                   10),
        ("Tram 5G",                   10),
        ("Repeater",                  10),
        ("Booster",                   10),
        ("Node truyen dan only",      20),
        ("Tram phu song TSCA",        18),
        ("Phan loai tram",            22),
        ("MORAN 3G",                  15),
        ("MORAN 4G",                  15),
        ("MORAN 5G",                  15),
        ("Ma PTM",                    14),
        ("Do cao dinh cot anten (m)", 22),
        ("Do cao cot anten (m)",      20),
        ("Dia chi",                   30),
        ("Ghi chu",                   30),
    ]

    wb, ws = _make_wb(headers)

    def b(val: bool) -> str:
        return "x" if val else ""

    for idx, s in enumerate(sites, start=1):
        row = idx + 1
        values = [
            idx,
            s.mien, s.tinh, s.phuong_xa, s.site_name_cu, s.site_name,
            s.site_vip, s.lat, s.long,
            b(s.tram_2g), b(s.tram_3g), b(s.tram_4g), b(s.tram_5g),
            b(s.repeater), b(s.booster),
            b(s.node_truyen_dan_only), b(s.tram_phu_song_tsca),
            s.phan_loai_tram,
            s.moran_3g, s.moran_4g, s.moran_5g, s.ma_ptm,
            s.do_cao_dinh_cot_anten, s.do_cao_cot_anten,
            s.dia_chi, s.ghi_chu,
        ]
        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=row, column=col_idx, value=val)
        _style_row(ws, row, len(headers), idx % 2 == 0)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    return _stream(wb, "Sites_Export.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# CELLS 3G export
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/cells-3g")
def export_cells_3g(
    search:        Optional[str] = Query(None),
    mien:          Optional[str] = Query(None),
    tinh:          Optional[str] = Query(None),
    vendor:        Optional[str] = Query(None),
    mimo:          Optional[str] = Query(None),
    vung_phu_song: Optional[str] = Query(None),
    db:    Session = Depends(get_db),
    _:     User    = Depends(get_optional_user),
):
    q = db.query(Cell3G)
    if search:
        q = q.filter(
            Cell3G.cell_name.ilike(f"%{search}%") |
            Cell3G.site_name.ilike(f"%{search}%")
        )
    if mien:          q = q.filter(Cell3G.mien == mien)
    if tinh:          q = q.filter(Cell3G.tinh == tinh)
    if vendor:        q = q.filter(Cell3G.vendor == vendor)
    if mimo:          q = q.filter(Cell3G.mimo == mimo)
    if vung_phu_song: q = q.filter(Cell3G.vung_phu_song == vung_phu_song)
    cells = q.order_by(Cell3G.mien, Cell3G.tinh, Cell3G.site_name,
                       Cell3G.cell_name).all()

    headers = [
        ("STT",            6), ("Mien",          8), ("Tinh",         22),
        ("Phuong xa",     22), ("Site Name",     25), ("Cell Name",    25),
        ("Cell VIP",      10), ("MORAN",         15), ("Lat",          14),
        ("Long",          14), ("Vung phu song", 15), ("Vendor",       14),
        ("Do cao anten",  15), ("Azimuth",       10), ("M-tilt",       10),
        ("E-Tilt",        10), ("Total Tilt",    12), ("Loai Anten",   30),
        ("Chung anten",   18), ("Baseband",      18), ("RF",           14),
        ("Cell ID",       14), ("ARFCN",         12), ("PSC",          10),
        ("MIMO",          10),
    ]

    wb, ws = _make_wb(headers)

    for idx, c in enumerate(cells, start=1):
        row = idx + 1
        values = [
            idx,
            c.mien, c.tinh, c.phuong_xa, c.site_name, c.cell_name,
            c.cell_vip, c.moran, c.lat, c.long,
            c.vung_phu_song, c.vendor, c.do_cao_anten,
            c.azimuth, c.m_tilt, c.e_tilt, c.total_tilt,
            c.loai_anten, c.chung_anten, c.baseband, c.rf,
            c.cell_id, c.arfcn, c.psc, c.mimo,
        ]
        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=row, column=col_idx, value=val)
        _style_row(ws, row, len(headers), idx % 2 == 0)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    return _stream(wb, "Cells_3G_Export.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# CELLS 4G export
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/cells-4g")
def export_cells_4g(
    search:        Optional[str] = Query(None),
    mien:          Optional[str] = Query(None),
    tinh:          Optional[str] = Query(None),
    vendor:        Optional[str] = Query(None),
    mimo:          Optional[str] = Query(None),
    vung_phu_song: Optional[str] = Query(None),
    db:    Session = Depends(get_db),
    _:     User    = Depends(get_optional_user),
):
    q = db.query(Cell4G)
    if search:
        q = q.filter(
            Cell4G.cell_name.ilike(f"%{search}%") |
            Cell4G.site_name.ilike(f"%{search}%")
        )
    if mien:          q = q.filter(Cell4G.mien == mien)
    if tinh:          q = q.filter(Cell4G.tinh == tinh)
    if vendor:        q = q.filter(Cell4G.vendor == vendor)
    if mimo:          q = q.filter(Cell4G.mimo == mimo)
    if vung_phu_song: q = q.filter(Cell4G.vung_phu_song == vung_phu_song)
    cells = q.order_by(Cell4G.mien, Cell4G.tinh, Cell4G.site_name,
                       Cell4G.cell_name).all()

    headers = [
        ("STT",              6), ("Mien",          8), ("Tinh",         22),
        ("Phuong xa",       22), ("Site Name",     25), ("Cell Name",    25),
        ("Cell VIP",        10), ("MORAN",         15), ("Lat",          14),
        ("Long",            14), ("Vung phu song", 15), ("Vendor",       14),
        ("Do cao anten",    15), ("Azimuth",       10), ("M-tilt",       10),
        ("E-Tilt",          10), ("Total Tilt",    12), ("Loai Anten",   30),
        ("Chung anten",     18), ("Baseband",      18), ("RF",           14),
        ("Cell ID",         14), ("EARFCN",        12), ("PCI",          10),
        ("Root Sequence ID",18), ("MIMO",          10),
    ]

    wb, ws = _make_wb(headers)

    for idx, c in enumerate(cells, start=1):
        row = idx + 1
        values = [
            idx,
            c.mien, c.tinh, c.phuong_xa, c.site_name, c.cell_name,
            c.cell_vip, c.moran, c.lat, c.long,
            c.vung_phu_song, c.vendor, c.do_cao_anten,
            c.azimuth, c.m_tilt, c.e_tilt, c.total_tilt,
            c.loai_anten, c.chung_anten, c.baseband, c.rf,
            c.cell_id, c.earfcn, c.pci, c.root_sequence_id, c.mimo,
        ]
        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=row, column=col_idx, value=val)
        _style_row(ws, row, len(headers), idx % 2 == 0)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    return _stream(wb, "Cells_4G_Export.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# CELLS 5G export
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/cells-5g")
def export_cells_5g(
    search:        Optional[str] = Query(None),
    mien:          Optional[str] = Query(None),
    tinh:          Optional[str] = Query(None),
    vendor:        Optional[str] = Query(None),
    mimo:          Optional[str] = Query(None),
    vung_phu_song: Optional[str] = Query(None),
    db:    Session = Depends(get_db),
    _:     User    = Depends(get_optional_user),
):
    q = db.query(Cell5G)
    if search:
        q = q.filter(
            Cell5G.cell_name.ilike(f"%{search}%") |
            Cell5G.site_name.ilike(f"%{search}%")
        )
    if mien:          q = q.filter(Cell5G.mien == mien)
    if tinh:          q = q.filter(Cell5G.tinh == tinh)
    if vendor:        q = q.filter(Cell5G.vendor == vendor)
    if mimo:          q = q.filter(Cell5G.mimo == mimo)
    if vung_phu_song: q = q.filter(Cell5G.vung_phu_song == vung_phu_song)
    cells = q.order_by(Cell5G.mien, Cell5G.tinh, Cell5G.site_name,
                       Cell5G.cell_name).all()

    headers = [
        ("STT",              6), ("Mien",          8), ("Tinh",         22),
        ("Phuong xa",       22), ("Site Name",     25), ("Cell Name",    25),
        ("Cell VIP",        10), ("MORAN",         15), ("Lat",          14),
        ("Long",            14), ("Vung phu song", 15), ("Vendor",       14),
        ("Do cao anten",    15), ("Azimuth",       10), ("M-tilt",       10),
        ("E-Tilt",          10), ("Total Tilt",    12), ("Loai Anten",   30),
        ("Baseband",        18), ("RF",            14), ("Cell ID",      14),
        ("NR-ARFCN",        12), ("PCI",           10),
        ("Root Sequence ID",18), ("MIMO",          10),
    ]

    wb, ws = _make_wb(headers)

    for idx, c in enumerate(cells, start=1):
        row = idx + 1
        values = [
            idx,
            c.mien, c.tinh, c.phuong_xa, c.site_name, c.cell_name,
            c.cell_vip, c.moran, c.lat, c.long,
            c.vung_phu_song, c.vendor, c.do_cao_anten,
            c.azimuth, c.m_tilt, c.e_tilt, c.total_tilt,
            c.loai_anten, c.baseband, c.rf,
            c.cell_id, c.nr_arfcn, c.pci, c.root_sequence_id, c.mimo,
        ]
        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=row, column=col_idx, value=val)
        _style_row(ws, row, len(headers), idx % 2 == 0)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    return _stream(wb, "Cells_5G_Export.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# ANTENNAS export
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/antennas")
def export_antennas(
    search: Optional[str] = Query(None),
    band:   Optional[str] = Query(None),
    db:     Session       = Depends(get_db),
    _:      User          = Depends(get_optional_user),
):
    q = db.query(Antenna)
    if search: q = q.filter(Antenna.name.ilike(f"%{search}%"))
    if band:   q = q.filter(Antenna.band.ilike(f"%{band}%"))
    antennas = q.order_by(Antenna.name).all()

    headers = [
        ("STT",            6), ("Name",           35), ("Band",          20),
        ("No of Ports",   12), ("No of Beam",     12), ("Horizontal BW", 14),
        ("Vertical BW",   12), ("Gain (dBi)",     12), ("Etilt range",   14),
        ("H (mm)",        10), ("W (mm)",         10), ("D (mm)",        10),
        ("Weight (kg)",   12), ("Connector type", 18), ("Ghi chu",       30),
    ]

    wb, ws = _make_wb(headers)

    for idx, a in enumerate(antennas, start=1):
        row = idx + 1
        values = [
            idx,
            a.name, a.band, a.no_of_ports, a.no_of_beam,
            a.horizontal_bw, a.vertical_bw, a.gain, a.etilt,
            a.h, a.w, a.d, a.weight, a.connector_type, a.ghi_chu,
        ]
        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=row, column=col_idx, value=val)
        _style_row(ws, row, len(headers), idx % 2 == 0)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    return _stream(wb, "Antennas_Export.xlsx")
