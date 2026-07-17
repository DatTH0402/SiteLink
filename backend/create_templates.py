"""
create_templates.py
-------------------
Generates Excel template files for import.
Run once: python create_templates.py
"""
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), "templates")
os.makedirs(TEMPLATE_DIR, exist_ok=True)

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(color="FFFFFF", bold=True, size=11)
REQUIRED_FILL = PatternFill("solid", fgColor="FFE699")
REQUIRED_FONT = Font(color="7B3F00", bold=True, size=11)
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN          = Side(style="thin", color="BFBFBF")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

VN_LAT_MIN, VN_LAT_MAX = 8.33,   23.39
VN_LON_MIN, VN_LON_MAX = 102.14, 109.47
AZI_MIN,    AZI_MAX    = 0,       359

MIEN_LIST     = ["MB", "MT", "MN"]
CELL_VIP_LIST = ["VIP", "VVIP"]
MORAN_LIST    = ["VNPT HOST", "MBF HOST"]
VENDOR_LIST   = ["Ericsson", "Nokia", "Huawei", "ZTE", "Samsung"]
VUNG_PHU_SONG = ["Indoor", "Outdoor"]
MIMO_LIST     = ["2x2", "4x4", "8x8"]
SITE_VIP_LIST = ["VIP", "VVIP"]
PHAN_LOAI_LIST = ["IBC", "Macro outdoor", "IBC + Outdoor", "Smallcell", "miniDAS"]
CHUNG_ANTEN_3G = ["3G", "3G/4G", "2G/3G/4G", "3G/4G/5G", "3G/5G"]
CHUNG_ANTEN_4G = ["4G", "2G/4G", "3G/4G", "2G/3G/4G", "4G/5G"]
BOOL_LIST      = ["x", ""]
MU_MIMO_LIST   = ["Yes", "No"]


def _dv_list(ws, col_letter, values, start_row=2, end_row=1000, error_msg="Vui lòng chọn từ danh sách"):
    joined = ",".join(values)
    dv = DataValidation(
        type="list", formula1=f'"{joined}"',
        allow_blank=True, showDropDown=False,
        showErrorMessage=True,
        errorTitle="Giá trị không hợp lệ", error=error_msg,
    )
    dv.sqref = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.add_data_validation(dv)


def _dv_decimal(ws, col_letter, min_val, max_val, start_row=2, end_row=1000, error_msg="Giá trị ngoài phạm vi"):
    dv = DataValidation(
        type="decimal", operator="between",
        formula1=str(min_val), formula2=str(max_val),
        allow_blank=True, showErrorMessage=True,
        errorStyle="warning", errorTitle="Giá trị ngoài phạm vi", error=error_msg,
    )
    dv.sqref = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.add_data_validation(dv)


def _dv_whole(ws, col_letter, min_val, max_val, start_row=2, end_row=1000, error_msg="Giá trị ngoài phạm vi"):
    dv = DataValidation(
        type="whole", operator="between",
        formula1=str(min_val), formula2=str(max_val),
        allow_blank=True, showErrorMessage=True,
        errorStyle="warning", errorTitle="Giá trị ngoài phạm vi", error=error_msg,
    )
    dv.sqref = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.add_data_validation(dv)


def style_header(ws, col_idx, value, required=False, width=20):
    cell = ws.cell(row=1, column=col_idx, value=value)
    cell.fill      = REQUIRED_FILL if required else HEADER_FILL
    cell.font      = REQUIRED_FONT if required else HEADER_FONT
    cell.alignment = CENTER
    cell.border    = BORDER
    ws.column_dimensions[get_column_letter(col_idx)].width = width


def add_example_row(ws, row_idx, values):
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.alignment = LEFT
        cell.border    = BORDER


def finalize(ws, num_cols):
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}1"


# ── SITE template ─────────────────────────────────────────────────────────────
def create_site_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sites"

    columns = [
        ("Mien",                         False, 10),
        ("Tinh",                         True,  22),
        ("Phuong xa",                    False, 22),
        ("Site name (cu)",               False, 22),
        ("Site name",                    True,  25),
        ("Site VIP",                     False, 12),
        ("Lat",                          False, 14),
        ("Long",                         False, 14),
        ("Tram 2G",                      False, 10),
        ("Tram 3G",                      False, 10),
        ("Tram 4G",                      False, 10),
        ("Tram 5G",                      False, 10),
        ("Repeater",                     False, 10),
        ("Booster",                      False, 10),
        ("Node truyen dan only",         False, 20),
        ("Tram phu song TSCA",           False, 18),
        ("Phan loai tram",               False, 22),
        ("MORAN 3G",                     False, 15),
        ("MORAN 4G",                     False, 15),
        ("MORAN 5G",                     False, 15),
        ("Ma PTM",                       False, 14),
        ("Do cao dinh cot anten",        False, 22),
        ("Do cao cot anten",             False, 20),
        ("Dia chi",                      False, 30),
        ("Ghi chu",                      False, 30),
    ]

    for i, (hdr, req, w) in enumerate(columns, start=1):
        style_header(ws, i, hdr, required=req, width=w)

    num_cols = len(columns)
    _dv_list(ws, "A", MIEN_LIST)
    _dv_list(ws, "F", SITE_VIP_LIST)
    _dv_decimal(ws, "G", VN_LAT_MIN, VN_LAT_MAX)
    _dv_decimal(ws, "H", VN_LON_MIN, VN_LON_MAX)
    for col_letter in ["I","J","K","L","M","N","O","P"]:
        _dv_list(ws, col_letter, BOOL_LIST)
    _dv_list(ws, "Q", PHAN_LOAI_LIST)
    for col_letter in ["R","S","T"]:
        _dv_list(ws, col_letter, MORAN_LIST)

    example = [
        "MB", "Ha Noi", "Phuong Trung Hoa", "HN-001-OLD",
        "HN-001", "VIP", 21.0285, 105.8542,
        "x", "x", "x", "x", "", "", "", "",
        "Macro outdoor", "MBF HOST", "MBF HOST", "",
        "PTM-001", 35.5, 30.0, "So 1, Duong ABC, Ha Noi", ""
    ]
    add_example_row(ws, 2, example)
    finalize(ws, num_cols)

    path = os.path.join(TEMPLATE_DIR, "template_site.xlsx")
    wb.save(path)
    print(f"  Created: {path}")


# ── CELL 3G template ──────────────────────────────────────────────────────────
# Required columns per spec:
# Baseband | RF | Cell ID | UARFCN | LAC | RAC | PSC | MIMO |
# URAId | Cell max power (dBm) | CPICH power (dBm) | BBUname | Cell status (at dump time)
def create_cell3g_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cells_3G"

    columns = [
        ("Mien",                       False, 10),   #  1  A
        ("Tinh",                       False, 22),   #  2  B
        ("Phuong xa",                  False, 22),   #  3  C
        ("Site Name Old",              False, 25),   #  4  D
        ("Cell Name Old",              False, 25),   #  5  E
        ("Site Name",                  True,  25),   #  6  F
        ("Cell Name",                  True,  25),   #  7  G
        ("Cell VIP",                   False, 12),   #  8  H
        ("MORAN",                      False, 15),   #  9  I
        ("Lat",                        False, 14),   # 10  J
        ("Long",                       False, 14),   # 11  K
        ("Vung phu song",              False, 15),   # 12  L
        ("Vendor",                     False, 14),   # 13  M
        ("Do cao anten",               False, 15),   # 14  N
        ("Azimuth",                    False, 12),   # 15  O
        ("M-tilt",                     False, 10),   # 16  P
        ("E-Tilt",                     False, 10),   # 17  Q
        ("Total Tilt",                 False, 12),   # 18  R
        ("Loai Anten",                 False, 30),   # 19  S
        ("Chung anten",                False, 18),   # 20  T
        ("Baseband",                   False, 18),   # 21  U
        ("RF",                         False, 14),   # 22  V
        ("Cell ID",                    False, 14),   # 23  W
        ("UARFCN",                     False, 12),   # 24  X
        ("LAC",                        False, 10),   # 25  Y
        ("RAC",                        False, 10),   # 26  Z
        ("PSC",                        False, 10),   # 27  AA
        ("MIMO",                       False, 10),   # 28  AB
        ("URAId",                      False, 10),   # 29  AC
        ("Cell max power (dBm)",       False, 20),   # 30  AD
        ("CPICH power (dBm)",          False, 18),   # 31  AE
        ("BBUname",                    False, 16),   # 32  AF
        ("Cell status (at dump time)", False, 24),   # 33  AG
    ]

    for i, (hdr, req, w) in enumerate(columns, start=1):
        style_header(ws, i, hdr, required=req, width=w)

    num_cols = len(columns)

    _dv_list(ws, "A", MIEN_LIST)
    _dv_list(ws, "H", CELL_VIP_LIST)
    _dv_list(ws, "I", MORAN_LIST)
    _dv_decimal(ws, "J", VN_LAT_MIN, VN_LAT_MAX)
    _dv_decimal(ws, "K", VN_LON_MIN, VN_LON_MAX)
    _dv_list(ws, "L", VUNG_PHU_SONG)
    _dv_list(ws, "M", VENDOR_LIST)
    _dv_whole(ws, "O", AZI_MIN, AZI_MAX)
    _dv_list(ws, "T", CHUNG_ANTEN_3G)
    _dv_list(ws, "AB", MIMO_LIST)

    example = [
        "MB", "Ha Noi", "Phuong Trung Hoa",
        "HN-001-OLD", "HN-001-3G-1-OLD",
        "HN-001", "HN-001-3G-1",
        "", "MBF HOST",
        21.0285, 105.8542, "Outdoor", "Huawei",
        30.0, 45, 2.0, 0.0, 2.0,
        "Huawei ATR4518R10v06", "3G", "BBU3910", "RRU3908",
        "12345",      # Cell ID
        "10562",      # UARFCN
        "1234",       # LAC
        "10",         # RAC
        "100",        # PSC
        "2x2",        # MIMO
        "1",          # URAId
        "43",         # Cell max power (dBm)
        "33",         # CPICH power (dBm)
        "BBU-HN-001", # BBUname
        "Active",     # Cell status
    ]
    add_example_row(ws, 2, example)
    finalize(ws, num_cols)

    path = os.path.join(TEMPLATE_DIR, "template_cell_3g.xlsx")
    wb.save(path)
    print(f"  Created: {path}")


# ── CELL 4G template ──────────────────────────────────────────────────────────
# Required columns per spec:
# Baseband | RF | EnodeB ID | Cell ID | EARFCN | TAC | PCI | Root Sequence ID |
# MIMO | Bandwidth | Cell max power (dBm) | ECI | BBUname | Cell status (at dump time)
def create_cell4g_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cells_4G"

    columns = [
        ("Mien",                       False, 10),   #  1  A
        ("Tinh",                       False, 22),   #  2  B
        ("Phuong xa",                  False, 22),   #  3  C
        ("Site Name Old",              False, 25),   #  4  D
        ("Cell Name Old",              False, 25),   #  5  E
        ("Site Name",                  True,  25),   #  6  F
        ("Cell Name",                  True,  25),   #  7  G
        ("Cell VIP",                   False, 12),   #  8  H
        ("MORAN",                      False, 15),   #  9  I
        ("Lat",                        False, 14),   # 10  J
        ("Long",                       False, 14),   # 11  K
        ("Vung phu song",              False, 15),   # 12  L
        ("Vendor",                     False, 14),   # 13  M
        ("Do cao anten",               False, 15),   # 14  N
        ("Azimuth",                    False, 12),   # 15  O
        ("M-tilt",                     False, 10),   # 16  P
        ("E-Tilt",                     False, 10),   # 17  Q
        ("Total Tilt",                 False, 12),   # 18  R
        ("Loai Anten",                 False, 30),   # 19  S
        ("Chung anten",                False, 18),   # 20  T
        ("Baseband",                   False, 18),   # 21  U
        ("RF",                         False, 14),   # 22  V
        ("EnodeB ID",                  False, 14),   # 23  W
        ("Cell ID",                    False, 14),   # 24  X
        ("EARFCN",                     False, 12),   # 25  Y
        ("TAC",                        False, 10),   # 26  Z
        ("PCI",                        False, 10),   # 27  AA
        ("Root Sequence ID",           False, 18),   # 28  AB
        ("MIMO",                       False, 10),   # 29  AC
        ("Bandwitdh",                  False, 12),   # 30  AD  (matches spec spelling)
        ("Cell max power (dBm)",       False, 20),   # 31  AE
        ("ECI",                        False, 12),   # 32  AF
        ("BBUname",                    False, 16),   # 33  AG
        ("Cell status (at dump time)", False, 24),   # 34  AH
    ]

    for i, (hdr, req, w) in enumerate(columns, start=1):
        style_header(ws, i, hdr, required=req, width=w)

    num_cols = len(columns)

    _dv_list(ws, "A", MIEN_LIST)
    _dv_list(ws, "H", CELL_VIP_LIST)
    _dv_list(ws, "I", MORAN_LIST)
    _dv_decimal(ws, "J", VN_LAT_MIN, VN_LAT_MAX)
    _dv_decimal(ws, "K", VN_LON_MIN, VN_LON_MAX)
    _dv_list(ws, "L", VUNG_PHU_SONG)
    _dv_list(ws, "M", VENDOR_LIST)
    _dv_whole(ws, "O", AZI_MIN, AZI_MAX)
    _dv_list(ws, "T", CHUNG_ANTEN_4G)
    _dv_list(ws, "AC", MIMO_LIST)

    example = [
        "MB", "Ha Noi", "Phuong Trung Hoa",
        "HN-001-OLD", "HN-001-4G-1-OLD",
        "HN-001", "HN-001-4G-1",
        "", "MBF HOST",
        21.0285, 105.8542, "Outdoor", "Huawei",
        30.0, 45, 2.0, 0.0, 2.0,
        "Huawei ATR4518R10v06", "4G", "BBU5900", "RRU5258",
        "12345",      # EnodeB ID
        "67890",      # Cell ID
        "1825",       # EARFCN
        "1234",       # TAC
        "100",        # PCI
        "0",          # Root Sequence ID
        "4x4",        # MIMO
        "20",         # Bandwidth
        "46",         # Cell max power (dBm)
        "1234567890", # ECI
        "BBU-HN-001", # BBUname
        "Active",     # Cell status
    ]
    add_example_row(ws, 2, example)
    finalize(ws, num_cols)

    path = os.path.join(TEMPLATE_DIR, "template_cell_4g.xlsx")
    wb.save(path)
    print(f"  Created: {path}")


# ── CELL 5G template ──────────────────────────────────────────────────────────
# Required columns per spec:
# Baseband | RF | gNodeB ID | Cell ID | TAC | PCI | Root Sequence ID | MIMO |
# SSB-ARFCN | Center-ARFCN | GSCN | Bandwidth (MHz) | Cell max power (dBm) |
# NCI | BBUname | MU-MIMO | Cell status (at dump time)
def create_cell5g_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cells_5G"

    columns = [
        ("Mien",                       False, 10),   #  1  A
        ("Tinh",                       False, 22),   #  2  B
        ("Phuong xa",                  False, 22),   #  3  C
        ("Site Name Old",              False, 25),   #  4  D
        ("Cell Name Old",              False, 25),   #  5  E
        ("Site Name",                  True,  25),   #  6  F
        ("Cell Name",                  True,  25),   #  7  G
        ("Cell VIP",                   False, 12),   #  8  H
        ("MORAN",                      False, 15),   #  9  I
        ("Lat",                        False, 14),   # 10  J
        ("Long",                       False, 14),   # 11  K
        ("Vung phu song",              False, 15),   # 12  L
        ("Vendor",                     False, 14),   # 13  M
        ("Do cao anten",               False, 15),   # 14  N
        ("Azimuth",                    False, 12),   # 15  O
        ("M-tilt",                     False, 10),   # 16  P
        ("E-Tilt",                     False, 10),   # 17  Q
        ("Total Tilt",                 False, 12),   # 18  R
        ("Loai Anten",                 False, 30),   # 19  S
        ("Baseband",                   False, 18),   # 20  T
        ("RF",                         False, 14),   # 21  U
        ("gNodeB ID",                  False, 14),   # 22  V
        ("Cell ID",                    False, 14),   # 23  W
        ("TAC",                        False, 10),   # 24  X
        ("PCI",                        False, 10),   # 25  Y
        ("Root Sequence ID",           False, 18),   # 26  Z
        ("MIMO",                       False, 10),   # 27  AA
        ("SSB-ARFCN",                  False, 12),   # 28  AB
        ("Center-ARFCN",               False, 14),   # 29  AC
        ("GSCN",                       False, 10),   # 30  AD
        ("Bandwidth (MHz)",            False, 14),   # 31  AE
        ("Cell max power (dBm)",       False, 20),   # 32  AF
        ("NCI",                        False, 12),   # 33  AG
        ("BBUname",                    False, 16),   # 34  AH
        ("MU-MIMO",                    False, 10),   # 35  AI
        ("Cell status (at dump time)", False, 24),   # 36  AJ
    ]

    for i, (hdr, req, w) in enumerate(columns, start=1):
        style_header(ws, i, hdr, required=req, width=w)

    num_cols = len(columns)

    _dv_list(ws, "A", MIEN_LIST)
    _dv_list(ws, "H", CELL_VIP_LIST)
    _dv_list(ws, "I", MORAN_LIST)
    _dv_decimal(ws, "J", VN_LAT_MIN, VN_LAT_MAX)
    _dv_decimal(ws, "K", VN_LON_MIN, VN_LON_MAX)
    _dv_list(ws, "L", VUNG_PHU_SONG)
    _dv_list(ws, "M", VENDOR_LIST)
    _dv_whole(ws, "O", AZI_MIN, AZI_MAX)
    _dv_list(ws, "AA", MIMO_LIST)
    _dv_list(ws, "AI", MU_MIMO_LIST)

    example = [
        "MB", "Ha Noi", "Phuong Trung Hoa",
        "HN-001-OLD", "HN-001-5G-1-OLD",
        "HN-001", "HN-001-5G-1",
        "", "MBF HOST",
        21.0285, 105.8542, "Outdoor", "Huawei",
        30.0, 45, 2.0, 0.0, 2.0,
        "Huawei AAU5614", "BBU5900", "AAU5614",
        "12345",       # gNodeB ID
        "11111",       # Cell ID
        "1234",        # TAC
        "100",         # PCI
        "0",           # Root Sequence ID
        "8x8",         # MIMO
        "627264",      # SSB-ARFCN
        "630048",      # Center-ARFCN
        "7999",        # GSCN
        "100",         # Bandwidth (MHz)
        "46",          # Cell max power (dBm)
        "123456789",   # NCI
        "BBU-HN-001",  # BBUname
        "Yes",         # MU-MIMO
        "Active",      # Cell status
    ]
    add_example_row(ws, 2, example)
    finalize(ws, num_cols)

    path = os.path.join(TEMPLATE_DIR, "template_cell_5g.xlsx")
    wb.save(path)
    print(f"  Created: {path}")


if __name__ == "__main__":
    create_site_template()
    create_cell3g_template()
    create_cell4g_template()
    create_cell5g_template()
    print("All templates created successfully.")
